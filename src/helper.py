import psutil
import time
import pandas as pd
from pathlib import Path
import os
from datetime import datetime
from dataclasses import dataclass
from openpyxl import load_workbook

import heatmap

def filter_non_empty_dfs(df):
    return df[df["DataFrames"].apply(lambda x: not x.empty)]


def export_results_to_dataframe(results, distance_type, obfuscation_type, dataset_type_first_part, dataset_type_last_part,
                                    split_algorithm_first_part, split_algorithm_last_part, seed, src_path):

    # Convert the results to a DataFrame
    results_df = pd.DataFrame(results)
    results_df = results_df.sort_values("Distance")

    # Calculate additional metrics
    yes_count = results_df['Equal File Name'].value_counts().get('Yes', 0)
    no_count = results_df['Equal File Name'].value_counts().get('No', 0)
    total_count = yes_count + no_count

    # Create a new DataFrame to store metrics
    metrics_df = pd.DataFrame({
        'Metric': ['Row_Count', 'Total_Yes', 'Total_No', 'Percentage_Yes', 'Percentage_No', 'Algorithm', 'Seed'],
        'Value': [
            len(results_df),
            yes_count,
            no_count,
            (yes_count / total_count) * 100 if total_count > 0 else 0,
            (no_count / total_count) * 100 if total_count > 0 else 0,
            distance_type,
            seed
        ]
    })

    # Append the metrics DataFrame to the final results DataFrame
    final_output_df = results_df.copy()
    final_output_df[''] = ''  # Create an empty column to separate data and metrics


    # Create output path and file name
    output_folder = Path('output/AP-Attack/')
    os.makedirs(output_folder, exist_ok=True)

    now = datetime.now()
    formatted_date = now.strftime("%Y%m%d_%H%M%S")
    output_file_name = (
        f"{obfuscation_type}_"
        f"{dataset_type_first_part if dataset_type_first_part == dataset_type_last_part else dataset_type_first_part + '-' + dataset_type_last_part}_"
        f"{split_algorithm_first_part if split_algorithm_first_part == split_algorithm_last_part else split_algorithm_first_part + '-' + split_algorithm_last_part}_"
        f"{formatted_date}_SEED{seed}.xlsx"
    )
    output_path = output_folder / output_file_name

    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        final_output_df.to_excel(writer, sheet_name='Results', index=False)
        metrics_df.to_excel(writer, sheet_name='Metrics', index=False)


def append_results_to_existing_file(results, distance_type, obfuscation_type, dataset_type_first_part, dataset_type_last_part,
                                    split_algorithm_first_part, split_algorithm_last_part, seed, src_path):
    # Generate output file name and formatted date
    formatted_date = pd.Timestamp.now().strftime("%Y%m%d")
    output_file_name = (
        f"{obfuscation_type}_"
        f"{dataset_type_first_part if dataset_type_first_part == dataset_type_last_part else dataset_type_first_part + '-' + dataset_type_last_part}_"
        f"{split_algorithm_first_part if split_algorithm_first_part == split_algorithm_last_part else split_algorithm_first_part + '-' + split_algorithm_last_part}_"
        f"{formatted_date}_SEED{seed}.xlsx"
    )

    # Convert the results to a DataFrame and sort it by Distance
    results_df = pd.DataFrame(results)
    results_df = results_df.sort_values("Distance")

    # Calculate additional metrics
    yes_count = results_df['Equal File Name'].value_counts().get('Yes', 0)
    no_count = results_df['Equal File Name'].value_counts().get('No', 0)
    total_count = yes_count + no_count

    # Create a new DataFrame to store metrics
    metrics_df = pd.DataFrame({
        'Metric': ['Row_Count', 'Total_Yes', 'Total_No', 'Percentage_Yes', 'Percentage_No', 'Algorithm', 'Seed'],
        'Value': [
            len(results_df),
            yes_count,
            no_count,
            (yes_count / total_count) * 100 if total_count > 0 else 0,
            (no_count / total_count) * 100 if total_count > 0 else 0,
            distance_type,
            seed
        ]
    })

    # Define the path for the results file
    output_folder = Path('output/AP-Attack/')
    os.makedirs(output_folder, exist_ok=True)
    results_file_path = output_folder / output_file_name
    
    # Writing or appending to the Excel file
    if os.path.exists(results_file_path):
        book = load_workbook(results_file_path)
        with pd.ExcelWriter(results_file_path, engine='openpyxl') as writer: 
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}

            # Determine starting rows
            startrow_results = writer.sheets['Results'].max_row + 2 if 'Results' in writer.sheets else 1
            startrow_metrics = writer.sheets['Metrics'].max_row + 2 if 'Metrics' in writer.sheets else 1
            
            # Insert header row in the Metrics sheet
            writer.sheets['Metrics'].cell(row=startrow_metrics, column=1, value=output_file_name)
            startrow_metrics += 1  # Shift the start row for the data

            # Write DataFrames to Excel sheets
            results_df.to_excel(writer, sheet_name='Results', index=False, startrow=startrow_results)
            metrics_df.to_excel(writer, sheet_name='Metrics', index=False, startrow=startrow_metrics)
    else:
        # If the file does not exist, create and write DataFrames with header row
        with pd.ExcelWriter(results_file_path, engine='openpyxl') as writer:
            # Write header and DataFrames
            worksheet_results = writer.book.create_sheet('Results')
            worksheet_metrics = writer.book.create_sheet('Metrics')
            
            # Write the header row
            worksheet_metrics.cell(row=1, column=1, value=output_file_name)

            # Write data to respective sheets
            results_df.to_excel(writer, sheet_name='Results', index=False, startrow=1)
            metrics_df.to_excel(writer, sheet_name='Metrics', index=False, startrow=2)

    print(f"Data has been successfully written to {results_file_path}")
    

def monitor_memory(stop_event):

    global current_memory_usage
    global memory_highest_value
    global memory_monitor_output
    
    while not stop_event.is_set():
        mem = psutil.virtual_memory()
        current_memory_usage = mem.percent

        if mem.percent > memory_highest_value:
            memory_highest_value = mem.percent

        if memory_monitor_output:
            print(f"Current memory usage: {mem.percent}, highest memory usage: {memory_highest_value}")
            
        time.sleep(1)  


def update_alteration_counter(counter, h, h_dash, u, v, reset):
    """
    During heat map alteration process, reset counter if topsoe divergence between H & V gets smaller (faster) compared to topsoe divergence between H & F. There exists
    a specified diff_previous value to prevent extremly long loops with barely any improvements
    """
    
    if reset:
        counter = 0 
    else:
        counter += 1 
    return counter


def precomputation(profiles: list, identifier_column: str) -> dict:
    """
    Precompute heatmaps and visited bins for a list of profiles
    
    Parameters:
        profiles (list of pd.DataFrame): A list of pandas DataFrames where each DataFrame corresponds to a profile
        identifier_column (str): The name of the column that contains the unique identifier for each profile
    
    Returns:
        dict: A dictionary where each key is the profile's identifier and the value is a ProfileData object
              containing the profile data, visited bins, and heatmap
    """
    
    # Define a data class to store profile data, visited bins, and heatmaps
    @dataclass
    class ProfileData:
        profile_data: pd.DataFrame  # Store the profile's data
        visited_bins: any  # Store the result of the visited bins calculation
        heat_map: any  # Store the generated heatmap 

    # Precompute and store the profile data, visited bins, and heatmaps for each profile
    precomputed_profiles = {
        profile.iloc[0][identifier_column]: ProfileData(
            profile_data=profile,
            visited_bins=heatmap.visited_bins(profile),  # Compute visited bins for the profile
            heat_map=heatmap.create_heatmap(profile)  # Generate a heatmap for the profile
        )
        for profile in profiles  # Iterate through each profile DataFrame in the list
    }

    # Return the dictionary of precomputed profiles
    return precomputed_profiles
